import json
import random
from datetime import datetime
import requests
from django.shortcuts import render, HttpResponse, redirect
from django.http import JsonResponse
from app01 import models
from app01.utils.pagelist import Pagination
from app01.utils.bootstrapmodelform import BootStrapModelForm, BootStrapForm
from app01.utils.encrypt import md5
from app01.utils.code import check_code
from io import BytesIO
import os
# 如果要用axaj发送post请求,导入csrf_exempt装饰器
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings


# Create your views here.
def depart_list(request):
    '''部门列表'''
    queryset = models.DepartMent.objects.all()
    return render(request, 'depart_list.html', locals())


def depart_add(request):
    '''新建部门'''
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    # 获取post提交过来的数据
    title = request.POST.get('title')
    # 存到数据库
    models.DepartMent.objects.create(title=title)
    return redirect('/depart/list/')


def depart_delete(request):
    '''删除部门'''
    if request.method == 'GET':
        # 获取id
        # http://127.0.0.1:8000/depart/delete/?nid=2
        nid = request.GET.get('nid')
        models.DepartMent.objects.filter(id=nid).delete()
        return redirect('/depart/list/')


# 另一种传参方式
def depart_edit(request, nid):
    '''修改部门'''

    # 根据id获取数据
    if request.method == 'GET':
        row_object = models.DepartMent.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', locals())
    title = request.POST.get('title')
    models.DepartMent.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')


def test(request):
    return render(request, 'test.html')


def user_list(request):
    '''用户列表'''
    userlist = models.UserInfo.objects.all()

    return render(request, 'user_list.html', locals())


#############ModelForm示例#############
from django import forms


class UserModelForm(forms.ModelForm):
    name = forms.CharField(min_length=2, label='用户名')  # 重写字段，控制字段输入长度

    class Meta:
        model = models.UserInfo
        fields = ['name', 'age', 'password', 'account', 'creat_time', 'depart', 'gender']
        # 添加一个插件，定义每个字段的表单属性
        widgets = {
            #     'name':forms.TextInput(attrs={"class":"form-control"}),
            "password": forms.PasswordInput(attrs={"class": "form-control", "placeholder": "请输入密码"})
            #     'age':forms.TextInput(attrs={'class':'form-control'}),
            #     'account': forms.TextInput(attrs={'class': 'form-control'}),
            #     'depart': forms.TextInput(attrs={'class': 'form-control'}),
            #     'gender': forms.TextInput(attrs={'class': 'form-control'}),
            #     'creat_time':forms.DateInput(attrs={'class':'form-control'})
        }

    # 批量定义字段表单属性
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            if name == 'password':
                continue
            field.widget.attrs = {'class': 'form-control', 'placeholder': f'请输入:{field.label}'}


def user_add(request):
    '''新建用户,基于modelform版本的'''
    if request.method == 'GET':
        form = UserModelForm()
        return render(request, 'user_add.html', locals())

    # 用户POST提交，数据校验
    form = UserModelForm(data=request.POST)
    if form.is_valid():
        # 如果数据合法，保存到数据库
        # print(form.cleaned_data)
        form.save()
        return redirect('/user/list/')
    else:
        return render(request, 'user_add.html', locals())
        print(form.errors)


def user_edit(request, nid):
    '''编辑用户'''
    row_object = models.UserInfo.objects.filter(id=nid).first()
    if request.method == 'GET':
        # 根据id获取行数据对象
        form = UserModelForm(instance=row_object)
        return render(request, 'user_edit.html', locals())
    form = UserModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        # 默认保存用户输入的值，如果需要自定义可以
        # form.instance.字段名 = 值
        form.save()
        return redirect('/user/list/')
    return render(request, 'user_edit.html', locals())


def user_delete(request, nid):
    '''删除用户'''
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect('/user/list/')


##################靓号管理####################
from django.core.validators import RegexValidator  # 正则验证
from django.core.exceptions import ValidationError  # 函数验证


class PerryNumForm(forms.ModelForm):
    # 设置手机号的正则表达式   验证手机号方式一
    # mobile = forms.CharField(
    #     label='手机号',
    #     validators=[RegexValidator(r'^1[3-9]\d{9}$','手机号格式错误')]
    # )
    class Meta:
        model = models.PrettyNum
        # fields = ['mobile','price','level','status'] #一般写法
        fields = "__all__"  # 所有的字段
        # exclude = ['level']  #排除level字段

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            if name == 'password':
                continue
            field.widget.attrs = {'class': 'form-control', 'placeholder': f'请输入:{field.label}'}

    # 验证手机号 方式二  函数命名 clean_字段名  ，可以验证是否存在之类的其他问题
    def clean_mobile(self):
        txt_mobile = self.cleaned_data["mobile"]
        exist = models.PrettyNum.objects.filter(mobile=txt_mobile).exists()  # 验证是否是已有数据
        if len(txt_mobile) != 11:
            raise ValidationError("格式错误！")
        if exist:
            raise ValidationError('手机号已存在！')
        return txt_mobile


def perttynum_list(request):
    '''靓号列表'''
    # 模拟数据
    # for i in range(300):
    #     models.PrettyNum.objects.create(mobile=f'1388888888{i}',price=88+i,level=1,status=1)

    # 查询结果
    data_dict = {}
    search_data = request.GET.get("q", "")

    if search_data:
        data_dict["mobile__contains"] = search_data

    queryset = models.PrettyNum.objects.filter(**data_dict).order_by('id')

    page_object = Pagination(request, queryset)

    page_queryset = page_object.page_queryset  # 分完页的数据
    page_string = page_object.html()  # 页码

    return render(request, 'perttynum_list.html', locals())


def perttynum_add(request):
    '''新建靓号'''
    if request.method == 'GET':
        form = PerryNumForm()
        return render(request, 'perttynum_add.html', locals())
    form = PerryNumForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/perttynum/list/')
    return render(request, 'perttynum_add.html', locals())


class PerryNumEditForm(forms.ModelForm):
    '''单独定义一个编辑的类'''
    # 设置手机号的正则表达式   验证手机号方式一
    # mobile = forms.CharField(
    #     label='手机号',
    #     validators=[RegexValidator(r'^1[3-9]\d{9}$','手机号格式错误')]
    # )
    mobile = forms.CharField(disabled=True, label='手机号')  # 设置手机号不可改

    class Meta:
        model = models.PrettyNum
        fields = ['mobile', 'level', 'status']  # 只显示部分字段
        # fields = "__all__"  #所有的字段
        # exclude = ['level']  #排除level字段

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, field in self.fields.items():
            # 字段中有属性则保留原来的属性，没属性加属性
            if field.widget.attrs:
                field.widget.attrs["class"] = 'form-control'
                field.widget.attrs["placeholder"] = f'请输入:{field.label}'
            else:
                field.widget.attrs = {'class': 'form-control', 'placeholder': f'请输入:{field.label}'}


def perttynum_edit(request, nid):
    '''编辑靓号'''
    row_object = models.PrettyNum.objects.filter(id=nid).first()
    if request.method == 'GET':
        form = PerryNumEditForm(instance=row_object)
        return render(request, 'perttynum_edit.html', locals())
    form = PerryNumEditForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/perttynum/list/')
    return render(request, 'perttynum_edit.html', locals())


def perttynum_delete(request, nid):
    '''删除靓号'''
    models.PrettyNum.objects.filter(id=nid).delete()
    return redirect('/perttynum/list/')


def admin_list(request):
    '''管理员列表'''
    info_dict = request.session["info"]

    qdata = models.Admin.objects.all()
    page_object = Pagination(request, qdata)
    queryset = page_object.page_queryset
    page_string = page_object.html()
    return render(request, 'admin_list.html', locals())


class AdminModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(label='确认密码', widget=forms.PasswordInput)

    class Meta:
        model = models.Admin
        fields = ['username', 'password', 'confirm_password']  # 只显示部分字段
        widgets = {'password': forms.PasswordInput}
        # fields = "__all__"  #所有的字段
        # exclude = ['level']  #排除level字段

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)

    def clean_confirm_password(self):
        confirm = md5(self.cleaned_data.get("confirm_password"))
        pwd = self.cleaned_data.get("password")
        if confirm != pwd:
            raise ValidationError("两次密码输入不一致！")
        return confirm


def admin_add(request):
    '''添加管理员'''
    title = '添加管理员'
    form = AdminModelForm()
    if request.method == 'GET':
        return render(request, 'change.html', locals())
    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    return render(request, 'change.html', locals())


class AdminEditModelForm(BootStrapModelForm):
    class Meta:
        model = models.Admin
        fields = ['username']


def admin_edit(request, nid):
    '''编辑管理员'''
    row_object = models.Admin.objects.filter(id=nid).first()
    title = '编辑管理员'
    error_msg = '用户不存在！'
    if not row_object:
        return render(request, 'error.html', locals())
    if request.method == 'GET':
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'change.html', locals())
    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    return render(request, 'change.html', locals())


def admin_delete(request, nid):
    '''删除管理员'''
    models.Admin.objects.filter(id=nid).delete()
    return redirect('/admin/list/')


class AdminResetModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(label='确认密码', widget=forms.PasswordInput)

    class Meta:
        model = models.Admin
        fields = ['password', 'confirm_password']  # 只显示部分字段
        widgets = {'password': forms.PasswordInput}
        # fields = "__all__"  #所有的字段
        # exclude = ['level']  #排除level字段

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        md5_pwd = md5(pwd)

        # 校验密码和新输入的是否一致
        exists = models.Admin.objects.filter(id=self.instance.pk, password=md5_pwd).exists()
        if exists:
            raise ValidationError("密码不能和之前的一样！")
        return md5_pwd

    def clean_confirm_password(self):
        confirm = md5(self.cleaned_data.get("confirm_password"))
        pwd = self.cleaned_data.get("password")
        if confirm != pwd:
            raise ValidationError("两次密码输入不一致！")
        return confirm


def admin_reset(request, nid):
    '''重置密码'''
    row_object = models.Admin.objects.filter(id=nid).first()
    title = f'重置---{row_object.username}---的密码'
    if not row_object:
        return redirect('/admin/list/')
    if request.method == 'GET':
        form = AdminResetModelForm(instance=row_object)
        return render(request, 'change.html', locals())
    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    return render(request, 'change.html', locals())


##########登陆##########
# 用Form的方式来实现
class LoginForm(BootStrapForm):
    username = forms.CharField(
        label="用户名",
        widget=forms.TextInput,
        required=True  # 必填
    )
    password = forms.CharField(
        label="密码",
        widget=forms.PasswordInput(render_value=True),
        required=True
    )
    code = forms.CharField(
        label='图片验证码',
        widget=forms.TextInput,
        required=True
    )

    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)


def login(request):
    '''用户登陆'''
    if request.method == "GET":
        form = LoginForm()
        return render(request, 'login.html', locals())
    form = LoginForm(data=request.POST)
    if form.is_valid():
        # 验证码的校验
        user_input_code = form.cleaned_data.pop('code')
        code = request.session.get('image_code', "")
        if code.upper() != user_input_code.upper():
            form.add_error("code", "验证码错误！")
            return render(request, 'login.html', locals())

        admin_object = models.Admin.objects.filter(**form.cleaned_data).first()
        if not admin_object:
            form.add_error("password", "用户名或密码错误！")
            return render(request, 'login.html', locals())
        request.session["info"] = {'id': admin_object.id, 'name': admin_object.username}
        request.session.set_expiry(60 * 60 * 24 * 7)  # 设置session过期时间为7天
        return redirect('/admin/list/')

    return render(request, 'login.html', locals())


def logout(request):
    '''注销'''
    request.session.clear()
    return redirect('/login/')


def img_code(request):
    '''生成验证码和图片'''
    img, code_string = check_code()

    # 将验证码存到session中
    request.session['image_code'] = code_string
    # 设置60秒超时
    request.session.set_expiry(60)
    stream = BytesIO()
    img.save(stream, 'png')
    return HttpResponse(stream.getvalue())


class TaskModelForm(BootStrapModelForm):
    class Meta:
        model = models.Task
        fields = "__all__"


@csrf_exempt
def task_list(request):
    '''任务管理'''
    queryset = models.Task.objects.all().order_by("-id")
    form = TaskModelForm()
    page_object = Pagination(request, queryset)

    page_queryset = page_object.page_queryset  # 分完页的数据
    page_string = page_object.html()  # 页码
    return render(request, 'task_list.html', locals())


@csrf_exempt
def task_ajax(request):
    # data_dict = {"status":True,"data":[11,22,33,44]}
    data_dict = {"status": True, "data": request.POST}
    return HttpResponse(json.dumps(data_dict))


@csrf_exempt
def task_add(request):
    # 1、用户发过来的数据进行校验（ModelForm验证）
    form = TaskModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        data_dict = {"status": True}
        return HttpResponse(json.dumps(data_dict))
    data_dict = {"status": False, "error": form.errors}
    return HttpResponse(json.dumps(data_dict, ensure_ascii=False))


class OrderModelForm(BootStrapModelForm):
    class Meta:
        model = models.Order
        exclude = ['oid', 'admin']  # 排除oid


def order_list(request):
    '''订单列表'''
    queryset = models.Order.objects.all().order_by("-id")
    form = OrderModelForm()

    page_object = Pagination(request, queryset)
    page_queryset = page_object.page_queryset  # 分完页的数据
    page_string = page_object.html()  # 页码

    return render(request, 'order_list.html', locals())


@csrf_exempt
def order_add(request):
    '''新建订单'''
    form = OrderModelForm(data=request.POST)
    if form.is_valid():
        form.instance.oid = datetime.now().strftime("%Y%m%d%H%M%S") + str(random.randint(1000, 9999))  # 自己生成订单编号
        # 获取当前登陆用户的id
        form.instance.admin_id = request.session["info"]["id"]  # ForeignKey一定有一个字段叫  XXX_id(可以在数据库中查出来)
        form.save()
        data_dict = {"status": True}
        return JsonResponse(data_dict)
    data_dict = {"status": False, "error": form.errors}
    return JsonResponse(data_dict)


def order_delete(request):
    '''删除订单'''
    uid = request.GET.get("uid")
    exists = models.Order.objects.filter(id=uid).exists()
    if not exists:
        data_dict = {"status": False, "eorror": "数据不存在！"}
        return JsonResponse(data_dict)
    data_dict = {"status": True}
    models.Order.objects.filter(id=uid).delete()
    return JsonResponse(data_dict)


def order_detail(request):
    '''编辑订单'''
    uid = request.GET.get("uid")
    row_dict = models.Order.objects.filter(id=uid).values("title", "price", "status").first()
    print(uid, row_dict)
    if not row_dict:
        data_dict = {"status": False, "eorror": "数据不存在！"}
        return JsonResponse(data_dict)
    data_dict = {"status": True, "data": row_dict}
    return JsonResponse(data_dict)


def chart_list(request):
    title = '数据统计'
    return render(request, 'chart_list.html', locals())


def chart_bar(request):
    '''柱状图'''
    legend = ['小波波', '阿肖肖']
    xAxis = ['一月', '二月', '三月', '四月', '五月', '六月']
    series = [
        {
            'name': '小波波',
            'type': 'bar',
            'data': [5, 20, 36, 10, 10, 20]
        },
        {
            'name': '阿肖肖',
            'type': 'bar',
            'data': [8, 31, 16, 15, 30, 10]
        }
    ]
    result = {
        'status': True,
        'data': {
            'legend': legend,
            'xAxis': xAxis,
            'series': series,
        }
    }
    return JsonResponse(result)


def chart_pie(request):
    '''饼状图'''
    result = {
        'status': True,
        'data': [
            {'value': 1048, 'name': 'IT部门'},
            {'value': 735, 'name': '新媒体'},
            {'value': 580, 'name': '销售部'},
        ]
    }
    return JsonResponse(result)

def chart_line(request):
    '''拆线图'''
    xAxis=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    series = [820, 932, 901, 934, 1290, 1330, 1320]
    result = {
        'status':True,
        'data':{
            'xAxis':xAxis,
            'series':series,
        }
    }
    return JsonResponse(result)

def upload_list(request):
    '''文件上传'''
    if request.method == "GET":
        return render(request,'upload_list.html')

    files_object = request.FILES.get('avatar')
    print(files_object.name)  #文件名
    f = open('a1.jpg',mode='wb')
    for chunk in files_object.chunks():
        f.write(chunk)
    f.close()
    return HttpResponse("...")

from openpyxl import load_workbook
def depart_muti(request):
    '''批量上传'''
    #1、获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    #2、使用openpyxl读取excel的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    #循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists=models.DepartMent.objects.filter(title=text).exists()
        if not exists:
            models.DepartMent.objects.create(title=text)

    return redirect('/depart/list/')

class UpForm(BootStrapForm):
    bootstrp_exclud_fields = ['img']
    name = forms.CharField(label='姓名')
    age = forms.IntegerField(label='年龄')
    img = forms.FileField(label='头像')

def upload_form(request):
    '''FORM上传'''
    title = 'Form上传'
    if request.method == "GET":
        form = UpForm()
        return render(request,'upload_form.html',locals())
    form = UpForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        #print(form.cleaned_data)
        #{'name': '老牛', 'age': 23, 'img': <InMemoryUploadedFile: logo.png (image/png)>}

        #1读取文件内容，写入到一个文件夹中
        img_object = form.cleaned_data.get('img')
        #media_path = os.path.join(settings.MEDIA_ROOT,img_object.name) #数据库中的文件路径，绝对路径
        media_path = os.path.join("media",img_object.name) #数据库中的文件路径，绝对路径
        f = open(media_path,mode='wb')
        for chunk in img_object.chunks():
            f.write(chunk)
        f.close()
        #2、将文件路径写入数据库中
        models.Boss.objects.create(
            name=form.cleaned_data['name'],
            age=form.cleaned_data['age'],
            img = media_path,
        )
        return HttpResponse('...')

class UploadModelForm(BootStrapModelForm):
    class Meta:
        model = models.City
        fields = "__all__"

def upload_modelform(request):
    '''基于modelform的上传'''
    title='基于modelform的上传'
    if request.method =="GET":
        form = UploadModelForm()
        return render(request,'upload_form.html',locals())  #共用upload_form.html
    form = UploadModelForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        form.save() #自动保存文件和存储数据库，文件存储的路径为models.py里设置的upload_to='city/'
        return HttpResponse("提交成功！")
    return render(request, 'upload_form.html', locals())  # 共用upload_form.html

def city_list(request):
    '''城市列表'''
    queryset = models.City.objects.all()
    return render(request,'city_list.html',locals())

class CityAddModelForm(BootStrapModelForm):
    class Meta:
        model = models.City
        fields = "__all__"
def city_add(request):
    '''添加城市'''
    title='添加城市'
    if request.method =="GET":
        form = CityAddModelForm()
        return render(request,'upload_form.html',locals())  #共用upload_form.html
    form = UploadModelForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        form.save() #自动保存文件和存储数据库，文件存储的路径为models.py里设置的upload_to='city/'
        return redirect('/city/list/')
    return render(request, 'upload_form.html', locals())  # 共用upload_form.html